import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font
import io
import zipfile
import re
import os
import sqlite3
from datetime import datetime, timedelta

# ==========================================
# CẤU HÌNH GIAO DIỆN & KHỞI TẠO CSDL
# ==========================================
st.set_page_config(page_title="🛠️ Công cụ Xử lý Dữ liệu Đào tạo VIAGS", layout="wide")

st.markdown("""
    <style>
    .main { background-color: #f8f9fa; }
    .stButton>button { width: 100%; border-radius: 5px; height: 3em; background-color: #007bff; color: white; }
    .status-box { padding: 20px; border-radius: 10px; border: 1px solid #dee2e6; background-color: white; }
    </style>
""", unsafe_allow_html=True)

DB_FILE = "dsnv_local.db"

def get_db_connection():
    """Tạo kết nối tới SQLite và trả về dạng Row để truy cập bằng tên cột"""
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Khởi tạo cấu trúc bảng CSDL - Chuẩn cấu trúc để đồng bộ Supabase sau này"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS nhan_vien (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ma_nv TEXT,
            ho_ten TEXT,
            don_vi_goc TEXT,
            don_vi_chuan TEXT
        )
    """)
    # Tạo Index cho ma_nv để tăng tốc độ truy vấn tối đa khi danh sách phình to
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_ma_nv ON nhan_vien (ma_nv)")
    conn.commit()
    conn.close()

# Chạy khởi tạo bảng ngay khi ứng dụng kích hoạt
init_db()

# ==========================================
# 📑 HÀM TIỆN ÍCH DÙNG CHUNG
# ==========================================
def remove_vietnamese_accents(s):
    s = str(s)
    s = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', s)
    s = re.sub(r'[ÈÉẸẺẼÊỀẾỆỂỄ]', 'E', s)
    s = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', s)
    s = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', s)
    s = re.sub(r'[ùúụủũưừứựửữ]', 'u', s)
    s = re.sub(r'[ìíịỉĩ]', 'i', s)
    s = re.sub(r'[ỳýỵỷỹ]', 'y', s)
    s = re.sub(r'[đ]', 'd', s)
    s = re.sub(r'[ÀÁẠẢÃÂẦẤẬẨẪĂẰẮẶ]', 'A', s)
    s = re.sub(r'[ÒÓỌỎÕÔỒỐỘỔỖƠỜỚỢỞỠ]', 'O', s)
    s = re.sub(r'[ÙÚỤỦŨƯỪỨỰỬỮ]', 'U', s)
    s = re.sub(r'[ÌÍỊỈĨ]', 'I', s)
    s = re.sub(r'[YÝỴỶỸ]', 'Y', s)
    s = re.sub(r'[Đ]', 'D', s)
    return s

def chuan_hoa_thoi_gian(tg_str):
    tg_str = str(tg_str).strip()
    tg_str = re.sub(r'^[CSTcst](?=\d)', '', tg_str) 
    tg_str = re.sub(r'\s*-\s*', ' - ', tg_str)
    return tg_str

def parse_date_range(thoi_gian_str):
    tg_clean = re.sub(r'^[CSTcst]', '', str(thoi_gian_str)).strip()
    tg_clean = tg_clean.replace('.', '/').replace('\u2013', '-').replace('\u2014', '-')
    parts = re.split(r'\s*(?:[–—-])\s*', tg_clean)

    full_dates = []
    day_only_parts = []
    for idx, part in enumerate(parts):
        part = part.strip()
        if not part:
            continue
        match_full = re.match(r'^(\d{1,2})\s*[\/\.-]\s*(\d{1,2})(?:\s*[\/\.-]\s*(\d{2,4}))?$', part)
        if match_full:
            day = int(match_full.group(1))
            month = int(match_full.group(2))
            year = int(match_full.group(3)) if match_full.group(3) else datetime.now().year
            if year < 100:
                year += 2000
            try:
                full_dates.append((idx, datetime(year, month, day)))
            except ValueError:
                continue
            continue
        match_day_only = re.match(r'^(\d{1,2})$', part)
        if match_day_only:
            day_only_parts.append((idx, int(match_day_only.group(1))))

    if full_dates:
        dates = [d for _, d in full_dates]
        for idx, day in day_only_parts:
            reference = next((d for j, d in full_dates if j > idx), None)
            if reference is None:
                reference = next((d for j, d in reversed(full_dates) if j < idx), None)
            if reference is None:
                reference = datetime.now()
            try:
                dates.append(datetime(reference.year, reference.month, day))
            except ValueError:
                continue
        dates.sort()
        return dates[0], dates[-1]

    if day_only_parts:
        day = day_only_parts[0][1]
        reference = datetime.now()
        try:
            date_val = datetime(reference.year, reference.month, day)
            return date_val, date_val
        except ValueError:
            return None, None

    return None, None

def week_labels_from_range(start_date, end_date):
    if not start_date or not end_date:
        return []
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    start_week = start_date - timedelta(days=start_date.weekday())
    week_num = start_week.isocalendar()[1]
    week_end = start_week + timedelta(days=6)
    
    if start_week.month == week_end.month:
        label = f"Tuần {week_num} ({start_week.day:02d}-{week_end.day:02d}/{start_week.month})"
    else:
        label = f"Tuần {week_num} ({start_week.day:02d}/{start_week.month}-{week_end.day:02d}/{week_end.month})"
    return [label]

def clean_header(val):
    return "".join(str(val).lower().split())

def read_excel_values_only(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith('.csv'):
        try: return {"CSV": pd.read_csv(uploaded_file, header=None, encoding='utf-8')}
        except:
            uploaded_file.seek(0)
            return {"CSV": pd.read_csv(uploaded_file, header=None, encoding='utf-8-sig')}
    elif name.endswith('.xls'):
        return pd.read_excel(uploaded_file, sheet_name=None, header=None, engine='xlrd')
    else:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        res = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data = [list(row) for row in ws.iter_rows(values_only=True)]
            res[sheet_name] = pd.DataFrame(data)
        return res

def chuan_hoa_don_vi(dv_str):
    if not dv_str or str(dv_str).lower() == 'nan': return ""
    dv_clean = re.sub(r'[\s\.\-\_]', '', str(dv_str).lower())
    dv_no_accent = remove_vietnamese_accents(dv_clean).lower() 
    
    mapping = {
        "hanhkhach": "PVHK", "pvhk": "PVHK", "hk": "PVHK",
        "hanghoa": "PVHH", "pvhh": "PVHH", "hh": "PVHH",
        "dieuhanh": "TTĐH", "ttdh": "TTĐH", "dh": "TTĐH", "đh": "TTĐH",
        "sando": "PVSĐ", "pvsd": "PVSĐ", "sd": "PVSĐ", "sđ": "PVSĐ",
        "hanhchinh": "KHHC", "khhc": "KHHC", "hc": "KHHC",
        "trung tâm huấn luyện": "VNBA", "Trung tâm Huấn luyện": "VNBA", "trungtamhuanluyen": "VNBA", "vnba": "VNBA",
        "ketoan": "KTOA", "ktoa": "KTOA", "kt": "KTOA",
        "chatluong": "QLCL", "qlcl": "QLCL", "cl": "QLCL"
    }
    for key, val in mapping.items():
        if key in dv_no_accent:
            return val
    return str(dv_str).strip().upper()

# ==========================================
# CORE LOGIC BÓC TÁCH KHĐT & TRỘN HỌC VIÊN
# ==========================================
def doc_khdt_gom_theo_tuan(file_khdt):
    file_khdt.seek(0)
    wb_khdt = pd.ExcelFile(file_khdt)
    
    GV_TTDT_MB = [
        "Ngô Bích Hằng", "Nguyễn Thị Phương Hiền", "Nguyễn Đăng Các", "Trần Tuấn Anh",
        "Đỗ Trí Thức", "Nguyễn Thu Hằng", "Ngô Trung Thành", "Nguyễn Hải Hà",
        "Vũ Hoàng Giang", "Đặng Khánh Ly", "Đỗ Thị Mỹ Bình", "Nguyễn Đức Nghĩa", "Nguyễn Thị Lan"
    ]
    
    sheet_vnba = next((n for n in wb_khdt.sheet_names if "vnba" in n.lower() or "nba" in n.lower()), None)
    sheet_vna = next((n for n in wb_khdt.sheet_names if "vna" in n.lower() and "vnba" not in n.lower()), None)
    
    target_sheets = []
    if sheet_vnba: target_sheets.append(("VNBA", sheet_vnba))
    if sheet_vna: target_sheets.append(("VNA", sheet_vna))
    
    if not target_sheets:
        valid_sheets = [n for n in wb_khdt.sheet_names if "đào tạo" in n.lower()]
        sheet_fallback = valid_sheets[0] if valid_sheets else wb_khdt.sheet_names[0]
        target_sheets.append(("VNBA", sheet_fallback))
        
    def is_roman(s): return bool(re.match(r'^(I{1,3}|IV|V|VI{1,3}|IX|X)$', str(s).strip(), re.IGNORECASE))
        
    classes, current_class = [], None
    dict_theo_tuan = {}
    
    df_list = []
    for sheet_type, sheet_name in target_sheets:
        df_raw = pd.read_excel(wb_khdt, sheet_name=sheet_name, header=None)
        header_row_idx = next((idx for idx, row in df_raw.iterrows() if any("khóa đào tạo" in str(s).lower() or "khoa dao tao" in str(s).lower() for s in row.values)), 4)
        
        df_data = pd.read_excel(wb_khdt, sheet_name=sheet_name, skiprows=header_row_idx)
        df_data["_SheetType"] = sheet_type
        df_list.append(df_data)
        
    df_khdt_data = pd.concat(df_list, ignore_index=True) if df_list else pd.DataFrame()
    # ==========================================
    # BẮT ĐẦU VÒNG LẶP QUÉT DỮ LIỆU
    # ==========================================
    current_roman_course = ""
    current_latin_course = "" 
    current_section_type = "OTHER"
    current_sheet_tracking = ""
    current_category = ""

    for _, row in df_khdt_data.iterrows():
        if len(row) < 12: continue
        sheet_type = row.get("_SheetType", "VNBA")
        
        if sheet_type != current_sheet_tracking:
            current_sheet_tracking = sheet_type
            current_roman_course = ""
            current_latin_course = ""
            current_section_type = "OTHER"
            current_category = ""
            
        stt, ten = str(row.iloc[0]).strip(), str(row.iloc[1]).strip()
        if stt.endswith('.0'): stt = stt[:-2]
            
        ten_lower_check = ten.lower()
        is_header_stt = is_roman(stt) or bool(re.match(r'^[A-Z](\.\d+)+$', stt))
        
        if is_header_stt:
            if "kiến thức chung" in ten_lower_check or "đào tạo bổ sung" in ten_lower_check:
                current_section_type = "KIEN_THUC"
                current_roman_course = ""
            elif "chuyên môn nghiệp vụ" in ten_lower_check:
                current_section_type = "CHUYEN_MON"
                current_roman_course = ""
            elif "nghiệp vụ" in ten_lower_check:
                current_section_type = "NGHIEP_VU"
                current_roman_course = ten
            else:
                current_section_type = "OTHER"
                current_roman_course = ten
            continue
            
        is_latin = bool(re.match(r'^\d+$', stt))
        if is_latin: current_latin_course = ten
            
        loai_hinh, el_val = str(row.iloc[2]).strip(), str(row.iloc[5]).strip()
        tu_ngay, den_ngay = str(row.iloc[8]).strip(), str(row.iloc[9]).strip()
        dia_diem, gv = str(row.iloc[10]).strip(), str(row.iloc[11]).strip()
        
        if pd.isna(row.iloc[1]) or ten == "" or ten == "nan" or "tổng cộng" in ten_lower_check: continue
        
        if sheet_type == "VNA":
            if gv and gv != "nan":
                is_ttdt_mb = any(gv_muc_tieu.lower() in gv.lower() for gv_muc_tieu in GV_TTDT_MB)
                if not is_ttdt_mb: continue  
            else:
                continue 

            ten_lower = ten.lower()
            if current_section_type == "NGHIEP_VU":
                ten = f"{ten}/{current_roman_course}"
            elif current_section_type == "KIEN_THUC":
                pass 
            else:
                if is_latin: pass
                else:
                    is_ly_thuyet = "lý thuyết" in ten_lower and "kiểm tra" in ten_lower
                    is_thuc_hanh = "nhóm" in ten_lower or stt == "-" or stt == "" or stt == "nan"
                    parent_course = current_latin_course if current_latin_course else current_roman_course
                    
                    if is_ly_thuyet: ten = f"Lý thuyết + kiểm tra/{parent_course}"
                    elif is_thuc_hanh: ten = f"Thực hành + kiểm tra/{parent_course}"
                    else: ten = f"{ten}/{parent_course}"
            
        ten_lower = ten.lower()
        if not stt.isdigit() and not is_roman(stt):
            if "đào tạo ban đầu" in ten_lower: current_category, current_roman_course = "Ban đầu", ""
            elif "đào tạo định kỳ" in ten_lower or "đào tạo nhắc lại" in ten_lower: current_category, current_roman_course = "Định kỳ", ""
            elif "đào tạo phục hồi" in ten_lower: current_category, current_roman_course = "Phục hồi", ""
            elif "bồi dưỡng" in ten_lower or "đào tạo bổ sung" in ten_lower or "đào tạo khác" in ten_lower: current_category, current_roman_course = "Bồi dưỡng kiến thức", ""
        
        if is_roman(stt) and stt != "" and stt != "nan":
            current_roman_course = ten; continue
            
        if stt.isdigit():
            if current_class: classes.append(current_class)
            final_loai_hinh = loai_hinh if loai_hinh != "nan" and loai_hinh != "" else current_category
            if final_loai_hinh.lower() == "bồi dưỡng kt" or final_loai_hinh.lower() == "bdkt":
                final_loai_hinh = "Bồi dưỡng kiến thức"
            has_elearning = el_val and el_val != "nan" and el_val.strip() != ""
            current_class = {
                "ten_goc": ten, "khoa_hoc_la_ma": current_roman_course, "loai_hinh": final_loai_hinh,
                "hinh_thuc": "Elearning + Trực tiếp" if has_elearning else "Trực tiếp",
                "dia_diem": dia_diem if dia_diem != "nan" else "", "dia_diem_lt": "", "dia_diem_th": "",
                "gv_lt": gv if gv != "nan" else "", "tu_ngay_goc": tu_ngay if tu_ngay != 'nan' else "",
                "den_ngay_goc": den_ngay if den_ngay != 'nan' else "", "lt_dates": [], "th_dates": [], "nhom_dates": []
            }
            if current_class["tu_ngay_goc"]: current_class["lt_dates"].append(current_class["tu_ngay_goc"])
            if current_class["den_ngay_goc"]: current_class["th_dates"].append(current_class["den_ngay_goc"])
                
        elif current_class is not None:
            row_text_lower = (stt + " " + ten).lower()
            if el_val and el_val != "nan" and el_val.strip() != "":
                current_class["hinh_thuc"] = "Elearning + Trực tiếp"
            
            if "nhóm" in row_text_lower:
                if tu_ngay != 'nan' and den_ngay != 'nan': current_class["nhom_dates"].append((tu_ngay, den_ngay))
            elif "lý thuyết" in row_text_lower or "lt" in row_text_lower.split() or "lý thuyết" in str(row.iloc[2]).lower():
                if tu_ngay != 'nan': current_class["lt_dates"].append(tu_ngay)
                if gv != 'nan' and not current_class["gv_lt"]: current_class["gv_lt"] = gv 
                if dia_diem != 'nan' and dia_diem: current_class["dia_diem_lt"] = dia_diem
            elif "thực hành" in row_text_lower or "th" in row_text_lower.split() or "thực hành" in str(row.iloc[2]).lower():
                if den_ngay != 'nan': current_class["th_dates"].append(den_ngay)
                elif tu_ngay != 'nan': current_class["th_dates"].append(tu_ngay)
                if dia_diem != 'nan' and dia_diem: current_class["dia_diem_th"] = dia_diem

    if current_class: classes.append(current_class)

    for c in classes:
        loai = str(c["loai_hinh"]).lower()
        ten_final = c["ten_goc"]
        thoi_gian = ""
        
        if "ban đầu" in loai or "ban dau" in loai:
            if c["khoa_hoc_la_ma"] and "tờ trình" not in c["khoa_hoc_la_ma"].lower():
                ten_final = f"{c['ten_goc']}/{c['khoa_hoc_la_ma']}"
            if c["nhom_dates"]:
                start, end = c["nhom_dates"][0][0], c["nhom_dates"][-1][1]
                thoi_gian = f"{start} - {end}" if start != end else start
            else:
                s, e = c["tu_ngay_goc"], c["den_ngay_goc"]
                thoi_gian = f"{s} - {e}" if s and e and s != e else (s or e)
            if not c["loai_hinh"]: c["loai_hinh"] = "Ban đầu"
        else:
            s = c["lt_dates"][0] if c["lt_dates"] else c["tu_ngay_goc"]
            e = c["th_dates"][-1] if c["th_dates"] else c["den_ngay_goc"]
            s_clean, e_clean = re.sub(r'^[CSTcst]', '', str(s)).strip(), re.sub(r'^[CSTcst]', '', str(e)).strip()
            if s_clean and e_clean and s_clean == e_clean: thoi_gian = s_clean
            elif s_clean and e_clean: thoi_gian = f"{s} - {e}" 
            else: thoi_gian = s_clean or e_clean
        
        if not thoi_gian: continue 
        
        dd_lt, dd_th = c["dia_diem_lt"] or c["dia_diem"], c["dia_diem_th"] or c["dia_diem"]
        final_dia_diem = f"{dd_lt} + {dd_th}" if dd_lt and dd_th and dd_lt != dd_th else (dd_lt or dd_th or c["dia_diem"])
        
        t_chuan = chuan_hoa_thoi_gian(thoi_gian)
        start_date, end_date = parse_date_range(t_chuan)
        week_labels = week_labels_from_range(start_date, end_date) if start_date and end_date else ["Tuần_Khác"]

        lop_info = {
            "ten_lop": ten_final, "thoi_gian": thoi_gian, "hinh_thuc": c["hinh_thuc"],
            "loai_hinh": c["loai_hinh"], "dia_diem": final_dia_diem, "giao_vien": c["gv_lt"]
        }

        for label in week_labels:
            if label not in dict_theo_tuan: dict_theo_tuan[label] = []
            dict_theo_tuan[label].append(lop_info)
        
    return dict_theo_tuan

def tao_file_excel_mot_tuan(tuan_name, dslop):
    wb = openpyxl.Workbook()
    ws_index = wb.active
    ws_index.title = "MucLuc"
    
    # Thiết lập độ rộng cột cho sheet Mục Lục rộng rãi để chứa trọn vẹn tên lớp dài
    ws_index.column_dimensions['A'].width = 8
    ws_index.column_dimensions['B'].width = 65  # TĂNG LÊN THEO YÊU CẦU CỦA BẠN
    ws_index.column_dimensions['C'].width = 25
    ws_index.column_dimensions['D'].width = 20
    ws_index.column_dimensions['E'].width = 25
    ws_index.column_dimensions['F'].width = 20
    
    headers = ["STT", "TÊN LỚP (BẤM VÀO ĐỂ TỚI SHEET)", "LOẠI HÌNH/HÌNH THỨC", "THỜI GIAN", "ĐỊA ĐIỂM", "GIÁO VIÊN"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_index.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name='Times New Roman', size=12, bold=True)
    
    for idx, lop in enumerate(dslop):
        stt = idx + 1
        row_idx = idx + 2
        ws_index[f'A{row_idx}'] = stt
        
        # [SỬA ĐỔI] Lấy luôn STT làm tên Sheet
        safe_name = str(stt)
            
        ws = wb.create_sheet(title=safe_name)
        
        # Tạo nút quay lại mục lục
        ws['A1'].value = '=HYPERLINK("#\'MucLuc\'!A1", "⬅️ Trở về Mục lục")'
        ws['A1'].font = Font(name='Times New Roman', size=12, color="0563C1", underline="single", bold=True)
        
        # Làm sạch tên lớp để hiển thị trong Mục Lục không bị lỗi
        clean_ten_lop = re.sub(r"[\n\r\t]", " ", str(lop["ten_lop"]))
        safe_display_name = clean_ten_lop.replace('"', '""') 
        
        # Ghi thông tin vào sheet Mục lục và tạo đường link trỏ sang sheet lớp (hiện tại tên là STT)
        link_cell = ws_index[f'B{row_idx}']
        link_cell.value = f'=HYPERLINK("#\'{safe_name}\'!A1", "{safe_display_name}")'
        link_cell.font = Font(name='Times New Roman', size=12, color="0563C1", underline="single")
        
        ws_index[f'C{row_idx}'] = f"{lop['loai_hinh']}/{lop['hinh_thuc']}"
        ws_index[f'D{row_idx}'] = lop["thoi_gian"]
        ws_index[f'E{row_idx}'] = lop["dia_diem"]
        ws_index[f'F{row_idx}'] = lop["giao_vien"]
        
        # Điền form thông tin khung lớp học
        ws['D7'], ws['D9'], ws['D10'], ws['D11'] = lop["ten_lop"], lop["thoi_gian"], lop["dia_diem"], lop["giao_vien"]
        ws['B7'] = f"- Môn học/Khóa học:"
        ws['B8'] = f"- Loại hình/hình thức đào tạo: {lop['loai_hinh']}/{lop['hinh_thuc']}"
        ws['B9'] = f"- Thời gian:"
        ws['B10'] = f"- Địa điểm:"
        ws['B11'] = f"- Giáo viên:"
        
        class_headers = ["STT", "Mã NV", "Họ tên", "Đơn vị", "Ghi chú"]
        for c_idx, h in enumerate(class_headers, start=1):
            ws.cell(row=13, column=c_idx, value=h).font = Font(name='Times New Roman', size=12, bold=True)
            
        for r_idx in range(14, 29):
            ws.cell(row=r_idx, column=1, value=r_idx-13)
            
    # ==========================================
    # QUÉT CHUẨN HÓA TOÀN BỘ FILE (FIX LỖI FILE & CĂN LỀ)
    # ==========================================
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Xử lý riêng cho sheet Mục Lục chính xác bằng so sánh chính xác tên viết liền
        if sheet_name.lower() == "mucluc":
            ws.column_dimensions['B'].width = 65  # Giữ cột B rộng rãi
            for row in ws.iter_rows(min_row=1):
                for cell in row:
                    if cell.value is not None:
                        # [QUAN TRỌNG] Né cột B ra để không làm mất màu xanh dương/gạch chân đặc trưng của link
                        if cell.column != 2:
                            is_bold = cell.font.bold if cell.font else False
                            cell.font = Font(name='Times New Roman', size=12, bold=is_bold)
        else:
            # Xử lý cho tất cả các sheet lớp học
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        if cell.coordinate != 'A1': # Giữ màu xanh cho nút quay lại ở ô A1
                            is_bold = cell.font.bold if cell.font else False
                            cell.font = Font(name='Times New Roman', size=12, bold=is_bold)
            
            # CĂN CHỈNH GIẢM ĐỘ RỘNG CÁC CỘT THEO YÊU CẦU (Rất gọn gàng khi in)
            ws.column_dimensions['A'].width = 6   # Cột STT
            ws.column_dimensions['B'].width = 14  # GIẢM XUỐNG THEO YÊU CẦU (Cột Mã NV)
            ws.column_dimensions['C'].width = 28  # Cột Họ tên
            ws.column_dimensions['D'].width = 14  # Cột Đơn vị
            ws.column_dimensions['E'].width = 15  # Cột Ghi chú
            
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

def tao_file_excel_tu_dict(dict_theo_tuan):
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for tuan, dslop in dict_theo_tuan.items():
            excel_buffer = tao_file_excel_mot_tuan(tuan, dslop)
            safe_tuan = str(tuan).replace("/", "-").replace(":", "")
            zf.writestr(f"Template_DSLop_{safe_tuan}.xlsx", excel_buffer.getvalue())
    zip_buffer.seek(0)
    return zip_buffer

def doc_dshv_ra_list(file_dshv):
    file_dshv.seek(0)
    wb_dshv = pd.ExcelFile(file_dshv)
    ds_lop_hv = []
    
    for sheetname in wb_dshv.sheet_names:
        s_name_clean = remove_vietnamese_accents(sheetname).lower()
        if s_name_clean in ["mucluc", "sheet1"]: continue
        
        if any(k in s_name_clean for k in ['ban dau', 'bandau', 'bsn', 'bsnd']): loai_hinh_sheet = 'Ban đầu'
        elif 'dinh ky' in s_name_clean or 'dinhky' in s_name_clean: loai_hinh_sheet = 'Định kỳ'
        elif 'phuc hoi' in s_name_clean or 'phuchoi' in s_name_clean: loai_hinh_sheet = 'Phục hồi'
        else: loai_hinh_sheet = 'Bồi dưỡng kiến thức' 
            
        df_sheet = pd.read_excel(wb_dshv, sheet_name=sheetname, header=None)
        header_row_hv = next((idx for idx, row in df_sheet.iterrows() if any("khóa học" in str(s).lower() or "khoa hoc" in str(s).lower() for s in row.values)), None)
        if header_row_hv is None: continue
        current_hv_class = None
        
        for idx in range(header_row_hv + 1, len(df_sheet)):
            row = df_sheet.iloc[idx]
            val_khoa_hoc = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            
            if val_khoa_hoc != "" and val_khoa_hoc != "nan" and not val_khoa_hoc.startswith("C."):
                tu_ngay_hv = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ""
                den_ngay_hv = str(row.iloc[7]).strip() if pd.notna(row.iloc[7]) else ""
                
                if tu_ngay_hv == "nan": tu_ngay_hv = ""
                if den_ngay_hv == "nan": den_ngay_hv = ""
                
                tg_hv = f"{tu_ngay_hv} - {den_ngay_hv}" if tu_ngay_hv != den_ngay_hv and den_ngay_hv else tu_ngay_hv
                ten_lop_hv = val_khoa_hoc.split("\n")[0].strip()
                
                current_hv_class = {"ten_lop": ten_lop_hv, "thoi_gian": tg_hv, "loai_hinh": loai_hinh_sheet, "hoc_vien": []}
                ds_lop_hv.append(current_hv_class)
                
            elif current_hv_class is not None:
                ma_nv = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ""
                ho_ten = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ""
                don_vi = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ""
                if ho_ten and ho_ten != "nan" and ho_ten != "HỌ VÀ TÊN":
                    current_hv_class["hoc_vien"].append({
                        "manv": ma_nv if ma_nv != "nan" else "", 
                        "hoten": ho_ten, 
                        "donvi": don_vi if don_vi != "nan" else ""
                    })
    return ds_lop_hv

def nhoi_hoc_vien_vao_template(file_template, ds_lop_hv):
    def norm_name(s):
        s = remove_vietnamese_accents(str(s)).lower()
        s = re.sub(r'[^a-z0-9]', ' ', s)
        return " ".join(s.split()) + " "

    def get_days(d_str):
        s = re.sub(r'/[0-9]{1,2}', '', str(d_str))
        nums = re.findall(r'\d+', s)
        return set([str(int(n)) for n in nums if int(n) <= 31])

    file_template.seek(0)
    wb = openpyxl.load_workbook(file_template)
    
    for sheetname in wb.sheetnames:
        if remove_vietnamese_accents(sheetname).lower() == "mucluc": continue
        ws = wb[sheetname]
        
        ten_lop_kh = str(ws['D7'].value or "")
        tg_kh = str(ws['D9'].value or "")
        b8_text = str(ws['B8'].value or "")
        
        loai_hinh_kh = ""
        m_loai = re.search(r'Loại hình:\s*([^/]+)', b8_text)
        if m_loai: loai_hinh_kh = m_loai.group(1).strip()
            
        key_kh_ten = norm_name(ten_lop_kh)
        kh_days = get_days(tg_kh)
        key_kh_loai = norm_name(loai_hinh_kh)[:6]
        
        matched_hv = []
        for hv_class in ds_lop_hv:
            key_hv_ten = norm_name(hv_class["ten_lop"])
            hv_days = get_days(hv_class["thoi_gian"])
            key_hv_loai = norm_name(hv_class["loai_hinh"])[:6]
            
            name_match = (key_kh_ten in key_hv_ten) or (key_hv_ten in key_kh_ten)
            type_match = (key_kh_loai == key_hv_loai)
            date_match = len(kh_days.intersection(hv_days)) > 0
            
            if name_match and type_match and (date_match or key_kh_ten == key_hv_ten):
                matched_hv = hv_class["hoc_vien"]
                break
                
        if matched_hv:
            for r in range(14, max(14 + len(matched_hv), 30)):
                for c in range(1, 6): ws.cell(row=r, column=c).value = ""
            for i, hv in enumerate(matched_hv):
                r_idx = 14 + i
                ws.cell(row=r_idx, column=1, value=i+1)
                ws.cell(row=r_idx, column=2, value=hv["manv"])
                ws.cell(row=r_idx, column=3, value=hv["hoten"])
                ws.cell(row=r_idx, column=4, value=hv["donvi"])

    out_buffer = io.BytesIO()
    wb.save(out_buffer)
    out_buffer.seek(0)
    return out_buffer

# ==========================================
# GIAO DIỆN CHÍNH (STREAMLIT TABS)
# ==========================================
st.title("🚀 Hệ thống Trộn & Đối chiếu Dữ liệu Đào tạo VIAGS")

tab_doi_chieu, tab_tao_khung, tab_nhoi_hv, tab_ql_csdl = st.tabs([
    "🔍 1. ĐỐI CHIẾU DANH SÁCH HỌC VIÊN", 
    "📄 2. TẠO KHUNG TỪ KHĐT", 
    "🧑‍🎓 3. TỰ ĐỘNG THÊM HỌC VIÊN",
    "🗄️ 4. QUẢN LÝ CƠ SỞ DỮ LIỆU"
])

# --- TAB 1: ĐỐI CHIẾU DANH SÁCH (TRUY VẤN REAL-TIME TỪ SQLITE) ---
with tab_doi_chieu:
    st.info("💡 Hướng dẫn: Tải lên file Học viên cần rà soát. Hệ thống tự động quét tìm chéo với Cơ sở dữ liệu nội bộ.")
    file_dshv = st.file_uploader("Chọn file Danh sách học viên cần kiểm tra", type=["xlsx", "xls", "csv"])

    if st.button("🚀 Bắt đầu quét đối chiếu", type="primary"):
        if not file_dshv:
            st.warning("⚠️ Vui lòng tải lên File Danh sách học viên cần kiểm tra!")
        else:
            with st.spinner("Đang truy vấn Cơ sở dữ liệu và phân tích dữ liệu..."):
                try:
                    # Kiểm tra CSDL có dữ liệu không trước khi chạy
                    conn = get_db_connection()
                    count_nv = conn.execute("SELECT COUNT(*) FROM nhan_vien").fetchone()[0]
                    if count_nv == 0:
                        st.error("⚠️ Cơ sở dữ liệu hiện đang TRỐNG! Vui lòng qua Tab 4 nạp file DSNV master trước.")
                        conn.close()
                        st.stop()

                    results = []
                    total_checked = 0
                    dict_hv_sheets = read_excel_values_only(file_dshv)
                    
                    kw_ma = ['mãnv', 'mnv', 'manv', 'mãsốnv', 'mãnhânviên', 'staffid']
                    kw_ten = ['họvàtên', 'họtên', 'fullname']
                    
                    for sheet_name, df_sheet in dict_hv_sheets.items():
                        c_ma, c_ten, c_dv = -1, -1, -1
                        current_class = "N/A"
                        pending_class_name = ""
                        
                        for idx, row in df_sheet.iterrows():
                            row_raw = [str(x).strip() if x is not None else "" for x in row.values]
                            row_cleaned = [clean_header(x) for x in row_raw]
                            
                            cell_0 = row_raw[0] 
                            cell_6 = row_raw[6] if len(row_raw) > 6 else "" 
                            cell_7 = row_raw[7] if len(row_raw) > 7 else "" 
                            
                            if cell_0 and "/" in str(cell_6):
                                current_class = f"{cell_0.split(chr(10))[0]} [{cell_6} - {cell_7}]"
                                pending_class_name = "" 
                            elif cell_0 and len(cell_0) > 10 and not any(k in clean_header(cell_0) for k in kw_ma + kw_ten):
                                pending_class_name = cell_0.split(chr(10))[0]
                                
                            if ("lý thuyết" in cell_0.lower() or "thực hành" in cell_0.lower()) and "/" in str(cell_6):
                                if pending_class_name:
                                    current_class = f"{pending_class_name} [{cell_6} - {cell_7}]"

                            tmp_ma = next((i for i, v in enumerate(row_cleaned) if any(k in v for k in kw_ma)), -1)
                            tmp_ten = next((i for i, v in enumerate(row_cleaned) if any(k in v for k in kw_ten)), -1)
                            tmp_dv = next((i for i, v in enumerate(row_cleaned) if ('trungtâm' in v or 'trungtam' in v) and 'đội' not in v and 'doi' not in v), -1)
                            
                            if tmp_ma != -1 and tmp_ten != -1:
                                c_ma, c_ten = tmp_ma, tmp_ten
                                if tmp_dv != -1: c_dv = tmp_dv
                                continue
                                
                            if c_ma != -1 and c_ten != -1:
                                ma_nv = row_raw[c_ma].replace('.0', '').strip()
                                if ma_nv and ma_nv.lower() not in ['none', 'nan', '']:
                                    ho_ten = row_raw[c_ten]
                                    total_checked += 1
                                    
                                    dv_hv_raw = row_raw[c_dv] if c_dv != -1 else (row_raw[5] if len(row_raw) > 5 else "")
                                    dv_hv_chuan = chuan_hoa_don_vi(dv_hv_raw)
                                    
                                    # --- TRUY VẤN REAL-TIME TỪ SQLITE ---
                                    # Lấy toàn bộ danh sách nhân viên khớp mã này (giải quyết triệt để trùng mã đa nhánh)
                                    cursor = conn.cursor()
                                    cursor.execute("SELECT ho_ten, don_vi_chuan FROM nhan_vien WHERE ma_nv = ?", (ma_nv,))
                                    candidates = cursor.fetchall()
                                    
                                    if candidates:
                                        best_match = None
                                        loi_msg_best = ["Sai họ tên", "Sai đơn vị"]
                                        
                                        for row_nv in candidates:
                                            ten_chuan = row_nv["ho_ten"]
                                            dv_chuan = row_nv["don_vi_chuan"]
                                            
                                            sai_ten = " ".join(ho_ten.lower().split()) != " ".join(ten_chuan.lower().split())
                                            sai_dv = (dv_chuan != "") and (dv_hv_chuan != dv_chuan)
                                            
                                            if not sai_ten and not sai_dv:
                                                loi_msg_best = []
                                                best_match = row_nv
                                                break
                                                
                                            current_loi = []
                                            if sai_ten: current_loi.append("Sai họ tên")
                                            if sai_dv: current_loi.append("Sai đơn vị")
                                            
                                            if len(current_loi) < len(loi_msg_best):
                                                loi_msg_best = current_loi
                                                best_match = row_nv
                                                
                                        if loi_msg_best:
                                            if best_match is None: best_match = candidates[0]
                                            results.append({
                                                "Lớp/Khóa": current_class, "Trang": sheet_name, "Mã NV": ma_nv,
                                                "Tên (File)": ho_ten, 
                                                "Tên đúng": best_match["ho_ten"] if "Sai họ tên" in loi_msg_best else "-",
                                                "Đơn vị (File)": dv_hv_chuan if "Sai đơn vị" in loi_msg_best else "-",
                                                "Đơn vị đúng": best_match["don_vi_chuan"] if "Sai đơn vị" in loi_msg_best else "-",
                                                "Lỗi": " & ".join(loi_msg_best)
                                            })
                                    else:
                                        results.append({
                                            "Lớp/Khóa": current_class, "Trang": sheet_name, "Mã NV": ma_nv,
                                            "Tên (File)": ho_ten, "Tên đúng": "❌ KHÔNG CÓ TRONG GỐC",
                                            "Đơn vị (File)": dv_hv_chuan, "Đơn vị đúng": "-", "Lỗi": "Mã NV lạ"
                                        })
                    conn.close()

                    st.divider()
                    st.subheader(f"📊 Kết quả kiểm tra (Tổng quét: {total_checked} HV)")
                    if not results:
                        st.success("🎉 Tuyệt vời! Danh sách khớp 100% với Cơ sở dữ liệu gốc.")
                    else:
                        df_res = pd.DataFrame(results)
                        st.error(f"Phát hiện {len(df_res)} lỗi cần chỉnh sửa.")
                        st.dataframe(df_res, use_container_width=True, hide_index=True)
                        
                        csv = df_res.to_csv(index=False).encode('utf-8-sig')
                        st.download_button("📥 Tải danh sách lỗi (.csv)", data=csv, file_name='loi_danh_sach.csv', mime='text/csv')

                except Exception as e:
                    st.error(f"❌ Lỗi xử lý dữ liệu: {str(e)}")

# --- TAB 2: CHỨC NĂNG TẠO KHUNG TUẦN TỪ KHĐT ---
with tab_tao_khung:
    st.info("💡 Bước 1: Upload KHĐT để tạo ra file Excel chuẩn (Có mục lục, khung rỗng sẵn sàng đón học viên).")
    file_khdt = st.file_uploader("📂 Chọn file Kế hoạch đào tạo (KHĐT)", type=["xlsx"])
    
    if file_khdt:
        if "dict_khdt" not in st.session_state or st.session_state.get("file_khdt_name") != file_khdt.name:
            with st.spinner("Đang phân tích dữ liệu KHĐT..."):
                st.session_state["dict_khdt"] = doc_khdt_gom_theo_tuan(file_khdt)
                st.session_state["file_khdt_name"] = file_khdt.name
        
        dict_khdt = st.session_state["dict_khdt"]
        danh_sach_tuan = list(dict_khdt.keys())
        
        chon_tat_ca = st.checkbox("Tạo toàn bộ các tuần (Xuất file ZIP)")
        tuan_duoc_chon = danh_sach_tuan if chon_tat_ca else st.multiselect("Hoặc tuỳ chọn các tuần cụ thể:", options=danh_sach_tuan, default=danh_sach_tuan[:1] if danh_sach_tuan else [])
            
        if st.button("🚀 Tạo Template rỗng", type="primary"):
            if tuan_duoc_chon:
                dict_chon = {k: v for k, v in dict_khdt.items() if k in tuan_duoc_chon}
                if len(dict_chon) == 1:
                    tuan_don = list(dict_chon.keys())[0]
                    excel_file = tao_file_excel_mot_tuan(tuan_don, dict_chon[tuan_don])
                    safe_tuan = str(tuan_don).replace("/", "-").replace(":", "")
                    st.download_button("📥 Tải Template Excel", data=excel_file, file_name=f"Template_DSLop_{safe_tuan}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                else:
                    zip_file = tao_file_excel_tu_dict(dict_chon)
                    st.download_button("📥 Tải Template ZIP", data=zip_file, file_name=f"Template_DSLop_Full.zip", mime="application/zip")
            else:
                st.warning("⚠️ Vui lòng chọn ít nhất 1 tuần!")

# --- TAB 3: CHỨC NĂNG NHỒI HỌC VIÊN TỰ ĐỘNG ---
with tab_nhoi_hv:
    st.info("💡 Tính năng đọc thông tin lớp ở ô D7 và D9 của file Khung để tự động bốc toàn bộ Học viên dán vào trang tương ứng.")
    col_x, col_y = st.columns(2)
    with col_x:
        file_template_in = st.file_uploader("📂 1. Chọn file Khung rỗng (File tuần đã tạo ở Bước 2)", type=["xlsx"], key="tpl_in")
    with col_y:
        file_dshv_in = st.file_uploader("📂 2. Chọn file Danh sách Học viên tổng (VD: File DS T6)", type=["xlsx"], key="dshv_in")
        
    if st.button("🪄 Bắt đầu khớp & Điền học viên", key="btn_nhoi", type="primary"):
        if file_template_in and file_dshv_in:
            with st.spinner("Đang dò tìm chéo dữ liệu Lớp và Học viên..."):
                ds_lop_hv = doc_dshv_ra_list(file_dshv_in)
                filled_excel = nhoi_hoc_vien_vao_template(file_template_in, ds_lop_hv)
                
                st.success("🎉 Khớp dữ liệu thành công!")
                st.download_button(
                    label="📥 Tải File Danh Sách Lớp Hoàn Chỉnh",
                    data=filled_excel,
                    file_name=file_template_in.name.replace("Template_", "DS_Lop_ChinhThuc_"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.warning("⚠️ Bạn cần tải lên đủ cả file Khung Tuần và file Danh Sách Học Viên tháng để thực hiện!")

# --- TAB 4: QUẢN LÝ CƠ SỞ DỮ LIỆU CHUYÊN NGHIỆP (MỚI THÊM) ---
with tab_ql_csdl:
    st.header("🗄️ Khu vực Quản trị Cơ sở dữ liệu Nhân viên")
    
    # Hiển thị nhanh số lượng bản ghi hiện có
    conn = get_db_connection()
    current_total = conn.execute("SELECT COUNT(*) FROM nhan_vien").fetchone()[0]
    
    col_metric1, col_metric2 = st.columns(2)
    with col_metric1:
        st.metric(label="👥 Tổng số Nhân viên trong CSDL hiện tại", value=f"{current_total:,} nhân sự")
    with col_metric2:
        st.success("💡 Cấu trúc bảng này chuẩn chỉnh tương thích Schema với PostgreSQL / Supabase sau này.")
        
    st.subheader("📥 Nạp / Làm mới Danh sách Nhân viên (Master File DSNV)")
    file_dsnv = st.file_uploader("Tải lên File Danh sách nhân viên toàn công ty mới nhất (.xlsx, .xls, .csv)", type=["xlsx", "xls", "csv"])
    
    if st.button("💾 Thực hiện Import/Cập nhật CSDL", type="primary"):
        if not file_dsnv:
            st.warning("⚠️ Vui lòng chọn file DSNV master trước khi bấm import!")
        else:
            with st.spinner("Đang phân tích và đối chiếu dữ liệu cũ - mới..."):
                try:
                    df_dsnv_all = read_excel_values_only(file_dsnv)
                    kw_ma = ['mãnv', 'mnv', 'manv', 'mãsốnv', 'mãnhânviên', 'staffid']
                    kw_ten = ['họvàtên', 'họtên', 'fullname']
                    kw_dv = ['phòngban', 'phòng', 'phongban', 'phong'] 

                    parsed_new_records = []
                    seen_records = set() # Chống lặp dữ liệu trong chính file Excel

                    # 1. Đọc và lọc sạch file Excel mới
                    for s_name, df_s in df_dsnv_all.items():
                        c_ma, c_ten, c_dv = -1, -1, -1
                        for idx, row in df_s.iterrows():
                            row_raw = [str(x).strip() if x is not None else "" for x in row.values]
                            row_cleaned = [clean_header(x) for x in row_raw]
                            
                            tmp_ma = next((i for i, v in enumerate(row_cleaned) if any(k in v for k in kw_ma)), -1)
                            tmp_ten = next((i for i, v in enumerate(row_cleaned) if any(k in v for k in kw_ten)), -1)
                            tmp_dv = next((i for i, v in enumerate(row_cleaned) if any(k in v for k in kw_dv)), -1)
                            
                            if tmp_ma != -1 and tmp_ten != -1:
                                c_ma, c_ten = tmp_ma, tmp_ten
                                if tmp_dv != -1: c_dv = tmp_dv
                                continue
                                
                            if c_ma != -1 and c_ten != -1:
                                m_val = row_raw[c_ma].replace('.0', '').strip()
                                t_val = row_raw[c_ten].strip()
                                dv_val_raw = row_raw[c_dv] if c_dv != -1 else ""
                                dv_chuan_hoa = chuan_hoa_don_vi(dv_val_raw)
                                
                                if m_val and m_val.lower() not in ['none', 'nan', '']:
                                    record_key = (m_val, t_val, dv_chuan_hoa)
                                    if record_key not in seen_records:
                                        seen_records.add(record_key)
                                        parsed_new_records.append({
                                            "ma_nv": m_val, "ho_ten": t_val,
                                            "don_vi_goc": dv_val_raw, "don_vi_chuan": dv_chuan_hoa
                                        })

                    if not parsed_new_records:
                        st.error("❌ Không tìm thấy dòng dữ liệu nhân sự hợp lệ nào từ file đã chọn.")
                    else:
                        conn = get_db_connection()
                        # 2. Rút toàn bộ dữ liệu CSDL cũ ra để so sánh
                        old_data = conn.execute("SELECT * FROM nhan_vien").fetchall()
                        old_dict = {}
                        for r in old_data:
                            m = r["ma_nv"]
                            if m not in old_dict: old_dict[m] = []
                            old_dict[m].append(dict(r))

                        insert_data = []
                        update_data = []
                        log_changes = []

                        # 3. Thuật toán so khớp (UPSERT Logic)
                        for new_r in parsed_new_records:
                            ma = new_r["ma_nv"]
                            ten = new_r["ho_ten"]
                            dv = new_r["don_vi_chuan"]
                            dv_goc = new_r["don_vi_goc"]

                            if ma not in old_dict:
                                # HOÀN TOÀN MỚI
                                insert_data.append((ma, ten, dv_goc, dv))
                                log_changes.append({"Mã NV": ma, "Họ tên": ten, "Đơn vị": dv, "Hành động": "✨ Thêm mới", "Chi tiết thay đổi": "Nhân sự mới"})
                            else:
                                old_records = old_dict[ma]
                                # Kiểm tra xem có khớp 100% không
                                exact_match = any(o["ho_ten"] == ten and o["don_vi_chuan"] == dv for o in old_records)
                                
                                if exact_match:
                                    continue # Bỏ qua, không làm gì cả để nhẹ hệ thống
                                
                                # CÓ SỰ THAY ĐỔI -> Đưa vào danh sách Cập nhật (Update)
                                best_old = old_records[0]
                                changes = []
                                if best_old["ho_ten"] != ten:
                                    changes.append(f"Tên: {best_old['ho_ten']} ➔ {ten}")
                                if best_old["don_vi_chuan"] != dv:
                                    changes.append(f"ĐV: {best_old['don_vi_chuan']} ➔ {dv}")
                                
                                chi_tiet = " | ".join(changes)
                                update_data.append((ten, dv_goc, dv, best_old["id"]))
                                log_changes.append({"Mã NV": ma, "Họ tên": ten, "Đơn vị": dv, "Hành động": "⚠️ Cập nhật", "Chi tiết thay đổi": chi_tiet})

                        # 4. Thực thi vào Database
                        if insert_data:
                            conn.executemany("INSERT INTO nhan_vien (ma_nv, ho_ten, don_vi_goc, don_vi_chuan) VALUES (?, ?, ?, ?)", insert_data)
                        if update_data:
                            conn.executemany("UPDATE nhan_vien SET ho_ten = ?, don_vi_goc = ?, don_vi_chuan = ? WHERE id = ?", update_data)
                        
                        conn.commit()
                        conn.close()

                        # 5. Hiển thị báo cáo Cảnh báo/Cập nhật cho người dùng
                        if log_changes:
                            st.success(f"🎉 Đồng bộ hoàn tất! Đã thêm mới {len(insert_data)} và Cập nhật {len(update_data)} nhân sự.")
                            df_log = pd.DataFrame(log_changes)
                            st.dataframe(df_log, use_container_width=True, hide_index=True)
                        else:
                            st.info("✅ Dữ liệu trong file Excel hoàn toàn khớp với CSDL hiện tại. Không có thay đổi nào được thực hiện.")

                except Exception as e:
                    st.error(f"❌ Lỗi trong quá trình nạp CSDL: {str(e)}")
                                        
    st.divider()
    st.subheader("🔍 Tìm kiếm nhanh nhân sự trong CSDL nội bộ")
    search_query = st.text_input("Nhập Mã nhân viên hoặc Tên cần tra cứu thử:")
    if search_query:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT ma_nv, ho_ten, don_vi_goc, don_vi_chuan FROM nhan_vien WHERE ma_nv LIKE ? OR ho_ten LIKE ? LIMIT 20",
            (f"%{search_query}%", f"%{search_query}%")
        )
        rows = cursor.fetchall()
        if rows:
            df_preview = pd.DataFrame([dict(r) for r in rows])
            st.dataframe(df_preview, use_container_width=True)
        else:
            st.info("Không tìm thấy kết quả khớp.")
            
    conn.close()